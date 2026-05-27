from __future__ import annotations

import csv
from dataclasses import dataclass, field
import hashlib
import json
import math
import os
import re
from pathlib import Path
import sys
from typing import Any, List, Sequence
import urllib.error
import urllib.parse
import urllib.request
from uuid import NAMESPACE_URL, uuid5
import zipfile
import xml.etree.ElementTree as ET


PROJECT_ROOT = Path(__file__).resolve().parents[1]
LOCAL_DEPS = PROJECT_ROOT / ".venv_deps"
if LOCAL_DEPS.exists() and str(LOCAL_DEPS) not in sys.path:
    sys.path.insert(0, str(LOCAL_DEPS))

TEXT_EXTENSIONS = {".md", ".txt", ".csv", ".tsv", ".docx", ".pdf", ".xlsx"}


def _normalize_text(text: str) -> str:
    return " ".join(text.split()).strip()


def _clean_document_text(text: str) -> str:
    cleaned = text.replace("\r\n", "\n").replace("\r", "\n").strip()
    cleaned = re.sub(r"\n{3,}", "\n\n", cleaned)
    return cleaned


def _load_dotenv() -> None:
    env_path = PROJECT_ROOT / ".env"
    if not env_path.exists():
        return

    for raw_line in env_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip("'").strip('"')
        if key and key not in os.environ:
            os.environ[key] = value


def _tokenize(text: str) -> List[str]:
    tokens: List[str] = []
    for match in re.finditer(r"[A-Za-z0-9_]+|[\u4e00-\u9fff]", text.lower()):
        token = match.group(0).strip()
        if token and token not in tokens:
            tokens.append(token)
    return tokens


def _read_text_file(path: Path) -> str:
    try:
        return path.read_text(encoding="utf-8")
    except UnicodeDecodeError:
        return path.read_text(encoding="utf-8", errors="replace")


def _read_delimited_table(path: Path, delimiter: str) -> str:
    rows: List[str] = []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle, delimiter=delimiter)
        for row in reader:
            values = [_normalize_text(value) for value in row]
            if any(values):
                rows.append(" | ".join(values))
    return "\n".join(rows)


def _read_xlsx(path: Path) -> str:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError("Reading .xlsx files requires openpyxl.") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    blocks: List[str] = []
    try:
        for sheet in workbook.worksheets:
            blocks.append(f"# Sheet: {sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                values = ["" if value is None else _normalize_text(str(value)) for value in row]
                if any(values):
                    blocks.append(" | ".join(values))
    finally:
        workbook.close()
    return "\n".join(blocks)


def _read_docx(path: Path) -> str:
    try:
        with zipfile.ZipFile(path) as archive:
            xml_bytes = archive.read("word/document.xml")
    except Exception as exc:
        raise RuntimeError(f"Could not read .docx file: {path}") from exc

    root = ET.fromstring(xml_bytes)
    namespace = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    blocks: List[str] = []
    for paragraph in root.findall(".//w:p", namespace):
        texts = [node.text or "" for node in paragraph.findall(".//w:t", namespace)]
        text = _normalize_text("".join(texts))
        if text:
            blocks.append(text)
    return "\n".join(blocks)


def _read_pdf(path: Path) -> str:
    try:
        from pypdf import PdfReader
    except Exception as exc:
        raise RuntimeError(
            "Reading .pdf files requires pypdf. Install it or keep using the project-local .venv_deps folder."
        ) from exc

    reader = PdfReader(str(path))
    pages: List[str] = []
    for index, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""
        text = _clean_document_text(text)
        if text:
            pages.append(f"# Page {index}\n{text}")
    return "\n\n".join(pages)


def read_document(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix in {".md", ".txt"}:
        return _read_text_file(path)
    if suffix == ".csv":
        return _read_delimited_table(path, ",")
    if suffix == ".tsv":
        return _read_delimited_table(path, "\t")
    if suffix == ".xlsx":
        return _read_xlsx(path)
    if suffix == ".docx":
        return _read_docx(path)
    if suffix == ".pdf":
        return _read_pdf(path)
    raise ValueError(f"Unsupported document type: {path.suffix}")


def chunk_text(text: str, chunk_size: int = 800, overlap: int = 120) -> List[str]:
    cleaned = _clean_document_text(text)
    if not cleaned:
        return []

    chunk_size = max(1, chunk_size)
    overlap = max(0, min(overlap, chunk_size - 1))

    chunks: List[str] = []
    start = 0
    length = len(cleaned)
    while start < length:
        end = min(length, start + chunk_size)
        chunk = cleaned[start:end].strip()
        if chunk:
            chunks.append(chunk)
        if end >= length:
            break
        start = max(end - overlap, start + 1)
    return chunks


@dataclass(slots=True)
class KnowledgeChunk:
    key: str
    chunk_id: str
    source: str
    text: str
    metadata: dict[str, Any] = field(default_factory=dict)

    def to_context_block(self) -> str:
        return f"[{self.source}#{self.chunk_id}]\n{self.text}"


@dataclass(slots=True)
class SearchHit:
    chunk: KnowledgeChunk
    score: float


class KnowledgeBase:
    """A tiny local RAG corpus with chunking and lexical retrieval."""

    def __init__(self, chunks: Sequence[KnowledgeChunk]) -> None:
        self.chunks = list(chunks)
        self._indexed_terms = {
            chunk.key: set(_tokenize(chunk.text + " " + chunk.source)) for chunk in self.chunks
        }

    @classmethod
    def from_directory(
        cls,
        root: Path,
        *,
        chunk_size: int = 800,
        overlap: int = 120,
    ) -> "KnowledgeBase":
        chunks: List[KnowledgeChunk] = []
        root = Path(root)
        if not root.exists():
            return cls([])

        for path in sorted(root.rglob("*")):
            if not path.is_file() or path.suffix.lower() not in TEXT_EXTENSIONS:
                continue
            text = read_document(path).strip()
            if not text:
                continue

            for index, chunk in enumerate(chunk_text(text, chunk_size=chunk_size, overlap=overlap), start=1):
                source = str(path.relative_to(root)).replace("\\", "/")
                chunks.append(
                    KnowledgeChunk(
                        key=f"{source}::{index:03d}",
                        chunk_id=f"{index:03d}",
                        source=source,
                        text=chunk,
                        metadata={"path": str(path)},
                    )
                )

        return cls(chunks)

    def search(self, query: str, top_k: int = 4) -> List[SearchHit]:
        query_tokens = _tokenize(query)
        if not query_tokens:
            return []

        scored: List[SearchHit] = []
        query_text = query.lower()
        for chunk in self.chunks:
            terms = self._indexed_terms.get(chunk.key, set())
            score = sum(1 for token in query_tokens if token in terms)
            if chunk.source.lower() in query_text:
                score += 1
            if any(token in chunk.text.lower() for token in query_tokens if len(token) > 1):
                score += 1
            if score:
                scored.append(SearchHit(chunk=chunk, score=score))

        scored.sort(key=lambda item: (-item.score, item.chunk.source, item.chunk.chunk_id))
        return scored[: max(1, top_k)]

    def format_context(self, query: str, top_k: int = 4) -> str:
        hits = self.search(query, top_k=top_k)
        if not hits:
            return "No relevant context found."
        blocks = [hit.chunk.to_context_block() for hit in hits]
        return "\n\n".join(blocks)

    def sources(self) -> List[str]:
        return sorted({chunk.source for chunk in self.chunks})

    def chunk_count(self) -> int:
        return len(self.chunks)


class HashEmbeddingModel:
    """Deterministic local embedding model for learning vector retrieval."""

    def __init__(self, dimensions: int = 384) -> None:
        self.dimensions = dimensions

    def embed(self, text: str) -> List[float]:
        vector = [0.0] * self.dimensions
        tokens = _tokenize(text)
        if not tokens:
            return vector

        for token in tokens:
            digest = hashlib.blake2b(token.encode("utf-8"), digest_size=8).digest()
            bucket = int.from_bytes(digest[:4], "big") % self.dimensions
            sign = 1.0 if digest[4] % 2 == 0 else -1.0
            vector[bucket] += sign

        norm = math.sqrt(sum(value * value for value in vector))
        if not norm:
            return vector
        return [value / norm for value in vector]


class QdrantVectorStore:
    def __init__(
        self,
        url: str | None = None,
        api_key: str | None = None,
        collection: str = "micro_agent_rag",
        dimensions: int = 384,
        timeout: float = 60.0,
    ) -> None:
        _load_dotenv()
        self.url = (url or os.getenv("QDRANT_API_URL", "")).rstrip("/")
        self.api_key = api_key or os.getenv("QDRANT_API_KEY", "")
        self.collection = collection
        self.dimensions = dimensions
        self.timeout = timeout
        if not self.url:
            raise RuntimeError("Missing QDRANT_API_URL in .env or environment.")
        if not self.api_key:
            raise RuntimeError("Missing QDRANT_API_KEY in .env or environment.")

    def recreate_collection(self) -> None:
        self._request("DELETE", f"/collections/{self.collection}", allow_404=True)
        self.ensure_collection()

    def ensure_collection(self) -> None:
        payload = {
            "vectors": {
                "size": self.dimensions,
                "distance": "Cosine",
            }
        }
        self._request("PUT", f"/collections/{self.collection}", payload)

    def upsert_chunks(self, chunks: Sequence[KnowledgeChunk], embeddings: Sequence[List[float]]) -> None:
        points = []
        for chunk, vector in zip(chunks, embeddings):
            point_id = str(uuid5(NAMESPACE_URL, chunk.key))
            points.append(
                {
                    "id": point_id,
                    "vector": vector,
                    "payload": {
                        "key": chunk.key,
                        "chunk_id": chunk.chunk_id,
                        "source": chunk.source,
                        "text": chunk.text,
                        "metadata": chunk.metadata,
                    },
                }
            )

        if not points:
            return
        self._request(
            "PUT",
            f"/collections/{self.collection}/points?wait=true",
            {"points": points},
        )

    def search(self, vector: List[float], limit: int = 4) -> List[SearchHit]:
        payload = {
            "vector": vector,
            "limit": max(1, limit),
            "with_payload": True,
        }
        data = self._request("POST", f"/collections/{self.collection}/points/search", payload)
        results = data.get("result", [])
        hits: List[SearchHit] = []
        for item in results:
            payload = item.get("payload") or {}
            chunk = KnowledgeChunk(
                key=str(payload.get("key", "")),
                chunk_id=str(payload.get("chunk_id", "")),
                source=str(payload.get("source", "")),
                text=str(payload.get("text", "")),
                metadata=payload.get("metadata") if isinstance(payload.get("metadata"), dict) else {},
            )
            hits.append(SearchHit(chunk=chunk, score=float(item.get("score", 0.0))))
        return hits

    def _request(
        self,
        method: str,
        path: str,
        payload: dict[str, Any] | None = None,
        *,
        allow_404: bool = False,
    ) -> dict[str, Any]:
        data = None if payload is None else json.dumps(payload).encode("utf-8")
        request = urllib.request.Request(
            f"{self.url}{path}",
            data=data,
            method=method,
            headers={
                "Content-Type": "application/json",
                "api-key": self.api_key,
                "Authorization": f"Bearer {self.api_key}",
            },
        )
        try:
            with urllib.request.urlopen(request, timeout=self.timeout) as response:
                body = response.read().decode("utf-8")
        except urllib.error.HTTPError as exc:
            if allow_404 and exc.code == 404:
                return {}
            body = exc.read().decode("utf-8", errors="ignore")
            raise RuntimeError(f"Qdrant API error: {exc.code} {body}") from exc
        except OSError as exc:
            raise RuntimeError(f"Qdrant network error: {exc}") from exc
        return json.loads(body) if body else {}


class QdrantKnowledgeBase:
    def __init__(
        self,
        vector_store: QdrantVectorStore | None = None,
        embedding_model: HashEmbeddingModel | None = None,
    ) -> None:
        self.embedding_model = embedding_model or HashEmbeddingModel()
        self.vector_store = vector_store or QdrantVectorStore(dimensions=self.embedding_model.dimensions)

    def index(self, knowledge_base: KnowledgeBase, *, recreate: bool = False) -> int:
        if recreate:
            self.vector_store.recreate_collection()
        else:
            self.vector_store.ensure_collection()

        embeddings = [self.embedding_model.embed(chunk.text) for chunk in knowledge_base.chunks]
        self.vector_store.upsert_chunks(knowledge_base.chunks, embeddings)
        return len(knowledge_base.chunks)

    def search(self, query: str, top_k: int = 4) -> List[SearchHit]:
        vector = self.embedding_model.embed(query)
        return self.vector_store.search(vector, limit=top_k)

    def format_context(self, query: str, top_k: int = 4) -> str:
        hits = self.search(query, top_k=top_k)
        if not hits:
            return "No relevant context found."
        return "\n\n".join(hit.chunk.to_context_block() for hit in hits)
